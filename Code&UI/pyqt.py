import Match as mt
import Search as search
import xlrd
import openpyxl
import sys
import copy
import random
from PyQt5 import QtCore, QtGui, QtWidgets
from MainWindow import Ui_MainWindow
from mode import Ui_Dialog as mode_dialog
from answer_right import Ui_answerright_dialog as answerright_dialog
from answer_movetoomuch import  Ui_Dialog as answermovetoomuch_dialog
from question import Ui_Dialog as question_dialog
from answer_wrong import Ui_Dialog as answerwrong_dialog
from uploadquestion import Ui_Dialog as uploadquestion_dialog
from upload_existed import Ui_Dialog as upload_existed_dialog
from upload_no_solution import Ui_Dialog as upload_no_solution_dialog
from upload_success import Ui_Dialog as upload_success_dialog
from produce_question import Ui_Dialog as produce_success_dialog
from level import Ui_Dialog as level_dialog


import sys
sys.setrecursionlimit(100000)
#   主窗口
class MainWindow(QtWidgets.QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(MainWindow, self).__init__()
        self.setupUi(self)
        # button
        self.mode_button.setStyleSheet("background-color: white")
        self.upload_question.setStyleSheet("background-color: white")
        self.submit_button.setStyleSheet("background-color: white")
        self.question_button.setStyleSheet("background-color: white")
        self.reset_button.setStyleSheet("background-color: white")
        self.showanswer_button.setStyleSheet("background-color: white")
        self.equal_generate.setStyleSheet("background-color: white")
        self.generate_new.setStyleSheet("background-color: white")
        # icon
        ico = QtGui.QIcon()
        ico.addPixmap(QtGui.QPixmap("ico.ico"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.setWindowIcon(ico)
        #   初始化其他几个对话框
        self.mode = mode_dialog()   #   选择模式
        self.question = question_dialog()
        self.answer_right = answerright_dialog()
        self.answer_wrong = answerwrong_dialog()
        self.answermovetoomuch = answermovetoomuch_dialog()
        self.uploadmyquestion = uploadquestion_dialog()
        self.upload_success = upload_success_dialog()
        self.upload_no_solution = upload_no_solution_dialog()
        self.upload_existed = upload_existed_dialog()
        self.produce_success = produce_success_dialog()
        self.level = level_dialog()
        #   初始查看答案和提交不能点击
        self.submit_button.setEnabled(False)
        self.showanswer_button.setEnabled(False)
        #   连接主界面上的按键点击
        #   1.选择模式的按键
        self.mode_button.clicked.connect(self.mode.show)
        #   2.选择关卡的按键
        self.question_button.clicked.connect(self.question_clicked)
        #   选择模式界面中的按键被点击
        self.mode.move_1.clicked.connect(self.move_1_clicked)
        self.mode.move_2.clicked.connect(self.move_2_clicked)
        #   选择关卡界面中的按键被点击
        self.question.pushButton.clicked.connect(self.question_confirm_clicked)
        #   3.我要出题按键被点击
        self.upload_question.clicked.connect(self.upload_question_clicked)
        self.uploadmyquestion.upload_submit.clicked.connect(self.upload_submit_clicked)
        self.uploadmyquestion.back.clicked.connect(self.upload_back_clicked)
        #   4.提交答案按键被点击
        self.submit_button.clicked.connect(self.submit_clicked)
        #   提交答案键里面的键被点击
        self.upload_existed.pushButton.clicked.connect(self.upload_existed.close)
        self.upload_no_solution.pushButton.clicked.connect(self.upload_no_solution.close)
        self.upload_success.pushButton.clicked.connect(self.upload_success.close)
        #   5.查看答案按键被点击
        self.showanswer_button.clicked.connect(self.showanswer_button_clicked)
        #   6.生成新问题按键被点击
        self.generate_new.clicked.connect(self.generate_new_clicked)
        self.produce_success.pushButton.clicked.connect(self.produce_success.close)
        #   7.重置按键被点击
        self.reset_button.clicked.connect(self.reset_button_clicked)
        #   8.生成新等式按键被点击
        self.equal_generate.clicked.connect(self.equal_generate_clicked)
        self.level.pushButton.clicked.connect(self.level.close)
        #   answermovetoomuch中的返回键被点击
        self.answermovetoomuch.pushButton.clicked.connect(self.answermovetoomuch.close)
        #   answerright中的确认键被点击
        self.answer_right.pushButton.clicked.connect(self.answer_right.close)
        #   answerwrong
        self.answer_wrong.pushButton.clicked.connect(self.answer_wrong.close)
        #   火柴棍被点击
        self.pos_0_1.clicked.connect(self.match_pos_0_1)
        self.pos_0_2.clicked.connect(self.match_pos_0_2)
        self.pos_0_3.clicked.connect(self.match_pos_0_3)
        self.pos_0_4.clicked.connect(self.match_pos_0_4)
        self.pos_0_5.clicked.connect(self.match_pos_0_5)
        self.pos_0_6.clicked.connect(self.match_pos_0_6)
        self.pos_0_7.clicked.connect(self.match_pos_0_7)
        self.pos_1_1.clicked.connect(self.match_pos_1_1)
        self.pos_1_2.clicked.connect(self.match_pos_1_2)
        self.pos_1_3.clicked.connect(self.match_pos_1_3)
        self.pos_1_4.clicked.connect(self.match_pos_1_4)
        self.pos_1_5.clicked.connect(self.match_pos_1_5)
        self.pos_1_6.clicked.connect(self.match_pos_1_6)
        self.pos_1_7.clicked.connect(self.match_pos_1_7)
        self.pos_3_1.clicked.connect(self.match_pos_3_1)
        self.pos_3_2.clicked.connect(self.match_pos_3_2)
        self.pos_3_3.clicked.connect(self.match_pos_3_3)
        self.pos_3_4.clicked.connect(self.match_pos_3_4)
        self.pos_3_5.clicked.connect(self.match_pos_3_5)
        self.pos_3_6.clicked.connect(self.match_pos_3_6)
        self.pos_3_7.clicked.connect(self.match_pos_3_7)
        self.pos_4_1.clicked.connect(self.match_pos_4_1)
        self.pos_4_2.clicked.connect(self.match_pos_4_2)
        self.pos_4_3.clicked.connect(self.match_pos_4_3)
        self.pos_4_4.clicked.connect(self.match_pos_4_4)
        self.pos_4_5.clicked.connect(self.match_pos_4_5)
        self.pos_4_6.clicked.connect(self.match_pos_4_6)
        self.pos_4_7.clicked.connect(self.match_pos_4_7)
        self.pos_5_1.clicked.connect(self.match_pos_5_1)
        self.pos_5_2.clicked.connect(self.match_pos_5_2)
        self.pos_5_3.clicked.connect(self.match_pos_5_3)
        self.pos_5_4.clicked.connect(self.match_pos_5_4)
        self.pos_5_5.clicked.connect(self.match_pos_5_5)
        self.pos_5_6.clicked.connect(self.match_pos_5_6)
        self.pos_5_7.clicked.connect(self.match_pos_5_7)
        self.pos_6_1.clicked.connect(self.match_pos_6_1)
        self.pos_6_2.clicked.connect(self.match_pos_6_2)
        self.pos_6_3.clicked.connect(self.match_pos_6_3)
        self.pos_6_4.clicked.connect(self.match_pos_6_4)
        self.pos_6_5.clicked.connect(self.match_pos_6_5)
        self.pos_6_6.clicked.connect(self.match_pos_6_6)
        self.pos_6_7.clicked.connect(self.match_pos_6_7)
        self.sig_0.clicked.connect(self.match_sig_1)
        self.sig_1.clicked.connect(self.match_sig_2)
        self.sig_3_4.clicked.connect(self.match_sig_3_4)
        self.pos_0_1.setStyleSheet("background-color: white")
        self.pos_0_2.setStyleSheet("background-color: white")
        self.pos_0_3.setStyleSheet("background-color: white")
        self.pos_0_4.setStyleSheet("background-color: white")
        self.pos_0_5.setStyleSheet("background-color: white")
        self.pos_0_6.setStyleSheet("background-color: white")
        self.pos_0_7.setStyleSheet("background-color: white")
        self.pos_1_1.setStyleSheet("background-color: white")
        self.pos_1_2.setStyleSheet("background-color: white")
        self.pos_1_3.setStyleSheet("background-color: white")
        self.pos_1_4.setStyleSheet("background-color: white")
        self.pos_1_5.setStyleSheet("background-color: white")
        self.pos_1_6.setStyleSheet("background-color: white")
        self.pos_1_7.setStyleSheet("background-color: white")
        self.pos_3_1.setStyleSheet("background-color: white")
        self.pos_3_2.setStyleSheet("background-color: white")
        self.pos_3_3.setStyleSheet("background-color: white")
        self.pos_3_4.setStyleSheet("background-color: white")
        self.pos_3_5.setStyleSheet("background-color: white")
        self.pos_3_6.setStyleSheet("background-color: white")
        self.pos_3_7.setStyleSheet("background-color: white")
        self.pos_4_1.setStyleSheet("background-color: white")
        self.pos_4_2.setStyleSheet("background-color: white")
        self.pos_4_3.setStyleSheet("background-color: white")
        self.pos_4_4.setStyleSheet("background-color: white")
        self.pos_4_5.setStyleSheet("background-color: white")
        self.pos_4_6.setStyleSheet("background-color: white")
        self.pos_4_7.setStyleSheet("background-color: white")
        self.pos_5_1.setStyleSheet("background-color: white")
        self.pos_5_2.setStyleSheet("background-color: white")
        self.pos_5_3.setStyleSheet("background-color: white")
        self.pos_5_4.setStyleSheet("background-color: white")
        self.pos_5_5.setStyleSheet("background-color: white")
        self.pos_5_6.setStyleSheet("background-color: white")
        self.pos_5_7.setStyleSheet("background-color: white")
        self.pos_6_1.setStyleSheet("background-color: white")
        self.pos_6_2.setStyleSheet("background-color: white")
        self.pos_6_3.setStyleSheet("background-color: white")
        self.pos_6_4.setStyleSheet("background-color: white")
        self.pos_6_5.setStyleSheet("background-color: white")
        self.pos_6_6.setStyleSheet("background-color: white")
        self.pos_6_7.setStyleSheet("background-color: white")
        self.sig_0.setStyleSheet("background-color: white")
        self.sig_1.setStyleSheet("background-color: white")
        self.sig_3_4.setStyleSheet("background-color: white")
        #   下面是初始化的一些变量值
        self.data_1 = xlrd.open_workbook('library_1.xlsx')
        self.workbook_1 = openpyxl.load_workbook('library_1.xlsx')
        self.data_2 = xlrd.open_workbook('library_2.xlsx')
        self.workbook_2 = openpyxl.load_workbook('library_2.xlsx')
        self.table_1 = self.data_1.sheets()[0]
        self.worksheet_1 = self.workbook_1.worksheets[0]
        self.worksheet_2 = self.workbook_2.worksheets[0]
        self.table_2 = self.data_2.sheets()[0]
        self.equal_initial = [copy.deepcopy(mt.num_empty),copy.deepcopy(mt.num_empty),copy.deepcopy(mt.sig_empty),copy.deepcopy(mt.num_empty),copy.deepcopy(mt.num_empty),
                         copy.deepcopy(mt.num_empty),copy.deepcopy(mt.num_empty)]
        #   可移动的火柴数
        self.match_movable = 1
        #   关卡序号
        self.question_number = 1
        #   游戏模式（默认为1）
        self.mode_number = 1
        #   两种模式的关卡数量
        self.question_number_mode1 = self.table_1.nrows - 1
        self.question_number_mode2 = self.table_2.nrows - 1
        #   可以选择的游戏关卡数目的上限值由当前的mode值决定
        self.question.spinBox.setMaximum(self.spinbox_set_maxvalue())
        #   定义全部的火柴棍flag
        self.flag_pos_0_1 = 1
        self.flag_pos_0_2 = 1
        self.flag_pos_0_3 = 1
        self.flag_pos_0_4 = 1
        self.flag_pos_0_5 = 1
        self.flag_pos_0_6 = 1
        self.flag_pos_0_7 = 1
        self.flag_pos_1_1 = 1
        self.flag_pos_1_2 = 1
        self.flag_pos_1_3 = 1
        self.flag_pos_1_4 = 1
        self.flag_pos_1_5 = 1
        self.flag_pos_1_6 = 1
        self.flag_pos_1_7 = 1
        self.flag_pos_3_1 = 1
        self.flag_pos_3_2 = 1
        self.flag_pos_3_3 = 1
        self.flag_pos_3_4 = 1
        self.flag_pos_3_5 = 1
        self.flag_pos_3_6 = 1
        self.flag_pos_3_7 = 1
        self.flag_pos_4_1 = 1
        self.flag_pos_4_2 = 1
        self.flag_pos_4_3 = 1
        self.flag_pos_4_4 = 1
        self.flag_pos_4_5 = 1
        self.flag_pos_4_6 = 1
        self.flag_pos_4_7 = 1
        self.flag_pos_5_1 = 1
        self.flag_pos_5_2 = 1
        self.flag_pos_5_3 = 1
        self.flag_pos_5_4 = 1
        self.flag_pos_5_5 = 1
        self.flag_pos_5_6 = 1
        self.flag_pos_5_7 = 1
        self.flag_pos_6_1 = 1
        self.flag_pos_6_2 = 1
        self.flag_pos_6_3 = 1
        self.flag_pos_6_4 = 1
        self.flag_pos_6_5 = 1
        self.flag_pos_6_6 = 1
        self.flag_pos_6_7 = 1
        #   符号的flag
        self.flag_sig_1 = 1
        self.flag_sig_2 = 1
        self.flag_sig_3 = 1
        self.flag_sig_4 = 1

        #
        #   设定火柴图片
        self.icon_match = QtGui.QIcon()
        self.icon_match.addPixmap(QtGui.QPixmap(":/newPrefix/Match.jpg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.icon_nothing = QtGui.QIcon()
        self.icon_nothing.addPixmap(QtGui.QPixmap(":/newPrefix/nothing.jpg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.icon_match_heng = QtGui.QIcon()
        self.icon_match_heng.addPixmap(QtGui.QPixmap("match_heng.jpg"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        #   按键全部禁用
        self.pos_0_1.setEnabled(False)
        self.pos_0_2.setEnabled(False)
        self.pos_0_3.setEnabled(False)
        self.pos_0_4.setEnabled(False)
        self.pos_0_5.setEnabled(False)
        self.pos_0_6.setEnabled(False)
        self.pos_0_7.setEnabled(False)
        self.pos_1_1.setEnabled(False)
        self.pos_1_2.setEnabled(False)
        self.pos_1_3.setEnabled(False)
        self.pos_1_4.setEnabled(False)
        self.pos_1_5.setEnabled(False)
        self.pos_1_6.setEnabled(False)
        self.pos_1_7.setEnabled(False)
        self.pos_3_1.setEnabled(False)
        self.pos_3_2.setEnabled(False)
        self.pos_3_3.setEnabled(False)
        self.pos_3_4.setEnabled(False)
        self.pos_3_5.setEnabled(False)
        self.pos_3_6.setEnabled(False)
        self.pos_3_7.setEnabled(False)
        self.pos_4_1.setEnabled(False)
        self.pos_4_2.setEnabled(False)
        self.pos_4_3.setEnabled(False)
        self.pos_4_4.setEnabled(False)
        self.pos_4_5.setEnabled(False)
        self.pos_4_6.setEnabled(False)
        self.pos_4_7.setEnabled(False)
        self.pos_5_1.setEnabled(False)
        self.pos_5_2.setEnabled(False)
        self.pos_5_3.setEnabled(False)
        self.pos_5_4.setEnabled(False)
        self.pos_5_5.setEnabled(False)
        self.pos_5_6.setEnabled(False)
        self.pos_5_7.setEnabled(False)
        self.pos_6_1.setEnabled(False)
        self.pos_6_2.setEnabled(False)
        self.pos_6_3.setEnabled(False)
        self.pos_6_4.setEnabled(False)
        self.pos_6_5.setEnabled(False)
        self.pos_6_6.setEnabled(False)
        self.pos_6_7.setEnabled(False)
        self.sig_0.setEnabled(False)
        self.sig_1.setEnabled(False)
        self.sig_3_4.setEnabled(False)
        self.pushButton_59.setEnabled(False)
        self.pushButton_60.setEnabled(False)
        self.submit_button.setEnabled(False)
    #   选择关卡键点击
    def question_clicked(self):
        self.question.spinBox.setMaximum(self.spinbox_set_maxvalue())
        self.question.show()
    #   在mode中如果移动一根被点击
    def move_1_clicked(self):
        self.mode.close()
        self.match_movable = 1
        self.mode_number = 1
        #   设定可以选择的关卡数目
        self.question.spinBox.setMaximum(self.question_number_mode1)
        self.question_number = 1
        self.output_question()
        self.pos_0_1.setEnabled(True)
        self.pos_0_2.setEnabled(True)
        self.pos_0_3.setEnabled(True)
        self.pos_0_4.setEnabled(True)
        self.pos_0_5.setEnabled(True)
        self.pos_0_6.setEnabled(True)
        self.pos_0_7.setEnabled(True)
        self.pos_1_1.setEnabled(True)
        self.pos_1_2.setEnabled(True)
        self.pos_1_3.setEnabled(True)
        self.pos_1_4.setEnabled(True)
        self.pos_1_5.setEnabled(True)
        self.pos_1_6.setEnabled(True)
        self.pos_1_7.setEnabled(True)
        self.pos_3_1.setEnabled(True)
        self.pos_3_2.setEnabled(True)
        self.pos_3_3.setEnabled(True)
        self.pos_3_4.setEnabled(True)
        self.pos_3_5.setEnabled(True)
        self.pos_3_6.setEnabled(True)
        self.pos_3_7.setEnabled(True)
        self.pos_4_1.setEnabled(True)
        self.pos_4_2.setEnabled(True)
        self.pos_4_3.setEnabled(True)
        self.pos_4_4.setEnabled(True)
        self.pos_4_5.setEnabled(True)
        self.pos_4_6.setEnabled(True)
        self.pos_4_7.setEnabled(True)
        self.pos_5_1.setEnabled(True)
        self.pos_5_2.setEnabled(True)
        self.pos_5_3.setEnabled(True)
        self.pos_5_4.setEnabled(True)
        self.pos_5_5.setEnabled(True)
        self.pos_5_6.setEnabled(True)
        self.pos_5_7.setEnabled(True)
        self.pos_6_1.setEnabled(True)
        self.pos_6_2.setEnabled(True)
        self.pos_6_3.setEnabled(True)
        self.pos_6_4.setEnabled(True)
        self.pos_6_5.setEnabled(True)
        self.pos_6_6.setEnabled(True)
        self.pos_6_7.setEnabled(True)
        self.sig_0.setEnabled(True)
        self.sig_1.setEnabled(True)
        self.sig_3_4.setEnabled(True)
        self.pushButton_59.setEnabled(True)
        self.pushButton_60.setEnabled(True)
        self.submit_button.setEnabled(True)
        self.showanswer_button.setEnabled(True)
    #   在mode中如果移动两根被点击
    def move_2_clicked(self):
        self.mode.close()
        self.match_movable = 2
        self.mode_number = 2
        self.question.spinBox.setMaximum(self.question_number_mode2)
        self.question_number = 1
        self.output_question()
        self.pos_0_1.setEnabled(True)
        self.pos_0_2.setEnabled(True)
        self.pos_0_3.setEnabled(True)
        self.pos_0_4.setEnabled(True)
        self.pos_0_5.setEnabled(True)
        self.pos_0_6.setEnabled(True)
        self.pos_0_7.setEnabled(True)
        self.pos_1_1.setEnabled(True)
        self.pos_1_2.setEnabled(True)
        self.pos_1_3.setEnabled(True)
        self.pos_1_4.setEnabled(True)
        self.pos_1_5.setEnabled(True)
        self.pos_1_6.setEnabled(True)
        self.pos_1_7.setEnabled(True)
        self.pos_3_1.setEnabled(True)
        self.pos_3_2.setEnabled(True)
        self.pos_3_3.setEnabled(True)
        self.pos_3_4.setEnabled(True)
        self.pos_3_5.setEnabled(True)
        self.pos_3_6.setEnabled(True)
        self.pos_3_7.setEnabled(True)
        self.pos_4_1.setEnabled(True)
        self.pos_4_2.setEnabled(True)
        self.pos_4_3.setEnabled(True)
        self.pos_4_4.setEnabled(True)
        self.pos_4_5.setEnabled(True)
        self.pos_4_6.setEnabled(True)
        self.pos_4_7.setEnabled(True)
        self.pos_5_1.setEnabled(True)
        self.pos_5_2.setEnabled(True)
        self.pos_5_3.setEnabled(True)
        self.pos_5_4.setEnabled(True)
        self.pos_5_5.setEnabled(True)
        self.pos_5_6.setEnabled(True)
        self.pos_5_7.setEnabled(True)
        self.pos_6_1.setEnabled(True)
        self.pos_6_2.setEnabled(True)
        self.pos_6_3.setEnabled(True)
        self.pos_6_4.setEnabled(True)
        self.pos_6_5.setEnabled(True)
        self.pos_6_6.setEnabled(True)
        self.pos_6_7.setEnabled(True)
        self.sig_0.setEnabled(True)
        self.sig_1.setEnabled(True)
        self.sig_3_4.setEnabled(True)
        self.pushButton_59.setEnabled(True)
        self.pushButton_60.setEnabled(True)
        self.submit_button.setEnabled(True)
        self.showanswer_button.setEnabled(True)
    #   选择关卡中可以选择的最大关卡数
    def spinbox_set_maxvalue(self):
        #   如果是模式一 则从移动一根的题库中找题
        if self.mode_number == 1:
            value_max = self.question_number_mode1
            return value_max
        else:
            value_max = self.question_number_mode2
            return value_max
    #   在选择关卡中点击确认键
    def question_confirm_clicked(self):
        self.question_number = self.question.spinBox.value()    #   value是当前的值
        self.question.close()
        #   根据question_number的值确定当前题目
        self.output_question()
        #   把按键全都解封
        self.pos_0_1.setEnabled(True)
        self.pos_0_2.setEnabled(True)
        self.pos_0_3.setEnabled(True)
        self.pos_0_4.setEnabled(True)
        self.pos_0_5.setEnabled(True)
        self.pos_0_6.setEnabled(True)
        self.pos_0_7.setEnabled(True)
        self.pos_1_1.setEnabled(True)
        self.pos_1_2.setEnabled(True)
        self.pos_1_3.setEnabled(True)
        self.pos_1_4.setEnabled(True)
        self.pos_1_5.setEnabled(True)
        self.pos_1_6.setEnabled(True)
        self.pos_1_7.setEnabled(True)
        self.pos_3_1.setEnabled(True)
        self.pos_3_2.setEnabled(True)
        self.pos_3_3.setEnabled(True)
        self.pos_3_4.setEnabled(True)
        self.pos_3_5.setEnabled(True)
        self.pos_3_6.setEnabled(True)
        self.pos_3_7.setEnabled(True)
        self.pos_4_1.setEnabled(True)
        self.pos_4_2.setEnabled(True)
        self.pos_4_3.setEnabled(True)
        self.pos_4_4.setEnabled(True)
        self.pos_4_5.setEnabled(True)
        self.pos_4_6.setEnabled(True)
        self.pos_4_7.setEnabled(True)
        self.pos_5_1.setEnabled(True)
        self.pos_5_2.setEnabled(True)
        self.pos_5_3.setEnabled(True)
        self.pos_5_4.setEnabled(True)
        self.pos_5_5.setEnabled(True)
        self.pos_5_6.setEnabled(True)
        self.pos_5_7.setEnabled(True)
        self.pos_6_1.setEnabled(True)
        self.pos_6_2.setEnabled(True)
        self.pos_6_3.setEnabled(True)
        self.pos_6_4.setEnabled(True)
        self.pos_6_5.setEnabled(True)
        self.pos_6_6.setEnabled(True)
        self.pos_6_7.setEnabled(True)
        self.sig_0.setEnabled(True)
        self.sig_1.setEnabled(True)
        self.sig_3_4.setEnabled(True)
        self.pushButton_59.setEnabled(True)
        self.pushButton_60.setEnabled(True)
        self.submit_button.setEnabled(True)
        self.showanswer_button.setEnabled(True)
    #   我要出题键被点击
    def upload_question_clicked(self):
        print(self.question_number_mode1)
        self.uploadmyquestion.show()
    #   出题的提交键被点击
    def upload_submit_clicked(self):
        equal = [copy.deepcopy(mt.num_empty), copy.deepcopy(mt.num_empty), copy.deepcopy(mt.sig_empty),
                         copy.deepcopy(mt.num_empty), copy.deepcopy(mt.num_empty),
                         copy.deepcopy(mt.num_empty), copy.deepcopy(mt.num_empty)]
        new_eq_str = self.uploadmyquestion.lineEdit.text()
        print(self.table_1.nrows)
        for i in range(self.question_number_mode1):
            if new_eq_str == self.worksheet_1.cell(i+2,1).value:
                self.upload_existed.show()
                self.uploadmyquestion.lineEdit.clear()
                return


        if mt.trans_string(new_eq_str, equal) == 1:
            outcome,equal = search.search_solution(equal)
            if outcome == 1:
                self.upload_success.show()
                self.worksheet_1.cell(self.question_number_mode1+2,1,new_eq_str)
                self.worksheet_1.cell(self.question_number_mode1+2,2,mt.equal2str(equal))
                self.workbook_1.save('library_1.xlsx')
                self.question_number_mode1 += 1
                self.data_1 = xlrd.open_workbook('library_1.xlsx')
                self.workbook_1 = openpyxl.load_workbook('library_1.xlsx')
                self.data_2 = xlrd.open_workbook('library_2.xlsx')
                self.workbook_2 = openpyxl.load_workbook('library_2.xlsx')
                self.table_1 = self.data_1.sheets()[0]
                self.worksheet_1 = self.workbook_1.worksheets[0]
                self.worksheet_2 = self.workbook_2.worksheets[0]
                self.table_2 = self.data_2.sheets()[0]
                self.question_number = self.question_number_mode1
                self.output_question()
                self.submit_button.setEnabled(True)
                self.showanswer_button.setEnabled(True)
                #   把按键全都解封
                self.pos_0_1.setEnabled(True)
                self.pos_0_2.setEnabled(True)
                self.pos_0_3.setEnabled(True)
                self.pos_0_4.setEnabled(True)
                self.pos_0_5.setEnabled(True)
                self.pos_0_6.setEnabled(True)
                self.pos_0_7.setEnabled(True)
                self.pos_1_1.setEnabled(True)
                self.pos_1_2.setEnabled(True)
                self.pos_1_3.setEnabled(True)
                self.pos_1_4.setEnabled(True)
                self.pos_1_5.setEnabled(True)
                self.pos_1_6.setEnabled(True)
                self.pos_1_7.setEnabled(True)
                self.pos_3_1.setEnabled(True)
                self.pos_3_2.setEnabled(True)
                self.pos_3_3.setEnabled(True)
                self.pos_3_4.setEnabled(True)
                self.pos_3_5.setEnabled(True)
                self.pos_3_6.setEnabled(True)
                self.pos_3_7.setEnabled(True)
                self.pos_4_1.setEnabled(True)
                self.pos_4_2.setEnabled(True)
                self.pos_4_3.setEnabled(True)
                self.pos_4_4.setEnabled(True)
                self.pos_4_5.setEnabled(True)
                self.pos_4_6.setEnabled(True)
                self.pos_4_7.setEnabled(True)
                self.pos_5_1.setEnabled(True)
                self.pos_5_2.setEnabled(True)
                self.pos_5_3.setEnabled(True)
                self.pos_5_4.setEnabled(True)
                self.pos_5_5.setEnabled(True)
                self.pos_5_6.setEnabled(True)
                self.pos_5_7.setEnabled(True)
                self.pos_6_1.setEnabled(True)
                self.pos_6_2.setEnabled(True)
                self.pos_6_3.setEnabled(True)
                self.pos_6_4.setEnabled(True)
                self.pos_6_5.setEnabled(True)
                self.pos_6_6.setEnabled(True)
                self.pos_6_7.setEnabled(True)
                self.sig_0.setEnabled(True)
                self.sig_1.setEnabled(True)
                self.sig_3_4.setEnabled(True)
                self.pushButton_59.setEnabled(True)
                self.pushButton_60.setEnabled(True)
                self.submit_button.setEnabled(True)
                self.showanswer_button.setEnabled(True)
            else:
                self.upload_no_solution.show()
                self.uploadmyquestion.lineEdit.clear()
        else:
            self.upload_no_solution.show()
            self.uploadmyquestion.lineEdit.clear()
    #   出题的返回键被点击
    #   清空lineedit,更新
    def upload_back_clicked(self):
        self.uploadmyquestion.close()
        self.uploadmyquestion.lineEdit.clear()
        self.spinbox_set_maxvalue()

    #   生成题目键被点击
    def generate_new_clicked(self):
        if self.produce_question() == 1:
            self.produce_success.show()
            #   把按键全都解封
            self.pos_0_1.setEnabled(True)
            self.pos_0_2.setEnabled(True)
            self.pos_0_3.setEnabled(True)
            self.pos_0_4.setEnabled(True)
            self.pos_0_5.setEnabled(True)
            self.pos_0_6.setEnabled(True)
            self.pos_0_7.setEnabled(True)
            self.pos_1_1.setEnabled(True)
            self.pos_1_2.setEnabled(True)
            self.pos_1_3.setEnabled(True)
            self.pos_1_4.setEnabled(True)
            self.pos_1_5.setEnabled(True)
            self.pos_1_6.setEnabled(True)
            self.pos_1_7.setEnabled(True)
            self.pos_3_1.setEnabled(True)
            self.pos_3_2.setEnabled(True)
            self.pos_3_3.setEnabled(True)
            self.pos_3_4.setEnabled(True)
            self.pos_3_5.setEnabled(True)
            self.pos_3_6.setEnabled(True)
            self.pos_3_7.setEnabled(True)
            self.pos_4_1.setEnabled(True)
            self.pos_4_2.setEnabled(True)
            self.pos_4_3.setEnabled(True)
            self.pos_4_4.setEnabled(True)
            self.pos_4_5.setEnabled(True)
            self.pos_4_6.setEnabled(True)
            self.pos_4_7.setEnabled(True)
            self.pos_5_1.setEnabled(True)
            self.pos_5_2.setEnabled(True)
            self.pos_5_3.setEnabled(True)
            self.pos_5_4.setEnabled(True)
            self.pos_5_5.setEnabled(True)
            self.pos_5_6.setEnabled(True)
            self.pos_5_7.setEnabled(True)
            self.pos_6_1.setEnabled(True)
            self.pos_6_2.setEnabled(True)
            self.pos_6_3.setEnabled(True)
            self.pos_6_4.setEnabled(True)
            self.pos_6_5.setEnabled(True)
            self.pos_6_6.setEnabled(True)
            self.pos_6_7.setEnabled(True)
            self.sig_0.setEnabled(True)
            self.sig_1.setEnabled(True)
            self.sig_3_4.setEnabled(True)
            self.pushButton_59.setEnabled(True)
            self.pushButton_60.setEnabled(True)
            self.submit_button.setEnabled(True)
            self.showanswer_button.setEnabled(True)
    #   重置键被点击
    def reset_button_clicked(self):
        self.output_question()
        #   把按键全都解封
        self.pos_0_1.setEnabled(True)
        self.pos_0_2.setEnabled(True)
        self.pos_0_3.setEnabled(True)
        self.pos_0_4.setEnabled(True)
        self.pos_0_5.setEnabled(True)
        self.pos_0_6.setEnabled(True)
        self.pos_0_7.setEnabled(True)
        self.pos_1_1.setEnabled(True)
        self.pos_1_2.setEnabled(True)
        self.pos_1_3.setEnabled(True)
        self.pos_1_4.setEnabled(True)
        self.pos_1_5.setEnabled(True)
        self.pos_1_6.setEnabled(True)
        self.pos_1_7.setEnabled(True)
        self.pos_3_1.setEnabled(True)
        self.pos_3_2.setEnabled(True)
        self.pos_3_3.setEnabled(True)
        self.pos_3_4.setEnabled(True)
        self.pos_3_5.setEnabled(True)
        self.pos_3_6.setEnabled(True)
        self.pos_3_7.setEnabled(True)
        self.pos_4_1.setEnabled(True)
        self.pos_4_2.setEnabled(True)
        self.pos_4_3.setEnabled(True)
        self.pos_4_4.setEnabled(True)
        self.pos_4_5.setEnabled(True)
        self.pos_4_6.setEnabled(True)
        self.pos_4_7.setEnabled(True)
        self.pos_5_1.setEnabled(True)
        self.pos_5_2.setEnabled(True)
        self.pos_5_3.setEnabled(True)
        self.pos_5_4.setEnabled(True)
        self.pos_5_5.setEnabled(True)
        self.pos_5_6.setEnabled(True)
        self.pos_5_7.setEnabled(True)
        self.pos_6_1.setEnabled(True)
        self.pos_6_2.setEnabled(True)
        self.pos_6_3.setEnabled(True)
        self.pos_6_4.setEnabled(True)
        self.pos_6_5.setEnabled(True)
        self.pos_6_6.setEnabled(True)
        self.pos_6_7.setEnabled(True)
        self.sig_0.setEnabled(True)
        self.sig_1.setEnabled(True)
        self.sig_3_4.setEnabled(True)
        self.pushButton_59.setEnabled(True)
        self.pushButton_60.setEnabled(True)
        self.submit_button.setEnabled(True)
    #   提交答案键被点击
    def submit_clicked(self):
        sticks_moved = self.sticks_moved()
        print(sticks_moved)
        #   先判断移动的火柴棍数是否正确
        if sticks_moved != 2 * self.match_movable:
            self.answermovetoomuch.show()
        #   移动的火柴棍数正确的情况下
        else:
            equal = self.read_equal()
            if mt.judge_equal(equal) == 1:
                self.answer_right.show()
            else:
                self.answer_wrong.show()
    #   显示答案按键被点击
    def showanswer_button_clicked(self):
        equal_initial = [copy.deepcopy(mt.num_empty), copy.deepcopy(mt.num_empty), copy.deepcopy(mt.sig_empty),
                          copy.deepcopy(mt.num_empty), copy.deepcopy(mt.num_empty),
                          copy.deepcopy(mt.num_empty), copy.deepcopy(mt.num_empty)]
        answer_initial = [copy.deepcopy(mt.num_empty),copy.deepcopy(mt.num_empty),copy.deepcopy(mt.sig_empty),copy.deepcopy(mt.num_empty),copy.deepcopy(mt.num_empty),
                         copy.deepcopy(mt.num_empty),copy.deepcopy(mt.num_empty)]
        if self.mode_number == 1:
            answer_str = self.table_1.cell(self.question_number, 1).value
        else:
            answer_str = self.table_2.cell(self.question_number, 1).value
        mt.trans_string(answer_str, answer_initial)
        if (answer_initial[2][2] == 1) & (answer_initial[2][3] == 1):
            self.equal_2.setCurrentIndex(1)
        else:
            self.equal_2.setCurrentIndex(0)
        self.set_match(answer_initial)
        #   然后令所有的button都不可点击，只能重新选择关卡
        self.pos_0_1.setEnabled(False)
        self.pos_0_2.setEnabled(False)
        self.pos_0_3.setEnabled(False)
        self.pos_0_4.setEnabled(False)
        self.pos_0_5.setEnabled(False)
        self.pos_0_6.setEnabled(False)
        self.pos_0_7.setEnabled(False)
        self.pos_1_1.setEnabled(False)
        self.pos_1_2.setEnabled(False)
        self.pos_1_3.setEnabled(False)
        self.pos_1_4.setEnabled(False)
        self.pos_1_5.setEnabled(False)
        self.pos_1_6.setEnabled(False)
        self.pos_1_7.setEnabled(False)
        self.pos_3_1.setEnabled(False)
        self.pos_3_2.setEnabled(False)
        self.pos_3_3.setEnabled(False)
        self.pos_3_4.setEnabled(False)
        self.pos_3_5.setEnabled(False)
        self.pos_3_6.setEnabled(False)
        self.pos_3_7.setEnabled(False)
        self.pos_4_1.setEnabled(False)
        self.pos_4_2.setEnabled(False)
        self.pos_4_3.setEnabled(False)
        self.pos_4_4.setEnabled(False)
        self.pos_4_5.setEnabled(False)
        self.pos_4_6.setEnabled(False)
        self.pos_4_7.setEnabled(False)
        self.pos_5_1.setEnabled(False)
        self.pos_5_2.setEnabled(False)
        self.pos_5_3.setEnabled(False)
        self.pos_5_4.setEnabled(False)
        self.pos_5_5.setEnabled(False)
        self.pos_5_6.setEnabled(False)
        self.pos_5_7.setEnabled(False)
        self.pos_6_1.setEnabled(False)
        self.pos_6_2.setEnabled(False)
        self.pos_6_3.setEnabled(False)
        self.pos_6_4.setEnabled(False)
        self.pos_6_5.setEnabled(False)
        self.pos_6_6.setEnabled(False)
        self.pos_6_7.setEnabled(False)
        self.sig_0.setEnabled(False)
        self.sig_1.setEnabled(False)
        self.sig_3_4.setEnabled(False)
        self.pushButton_59.setEnabled(False)
        self.pushButton_60.setEnabled(False)
        self.submit_button.setEnabled(False)



    #   下面是一些逻辑函数
    #   定义一个根据equal来设置按键火柴的函数
    def set_match(self, equal):
        pos_0 = equal[0]
        pos_1 = equal[1]
        pos_2 = equal[2]
        pos_3 = equal[3]
        pos_4 = equal[4]
        pos_5 = equal[5]
        pos_6 = equal[6]
        #   第一个数字
        if pos_0[0] == 1:
            self.flag_pos_0_1 = 1
            self.pos_0_1.setIcon(self.icon_match_heng)
        if pos_0[0] == 0:
            self.flag_pos_0_1 = 0
            self.pos_0_1.setIcon(self.icon_nothing)
        if pos_0[1] == 1:
            self.flag_pos_0_2 = 1
            self.pos_0_2.setIcon(self.icon_match)
        if pos_0[1] == 0:
            self.flag_pos_0_2 = 0
            self.pos_0_2.setIcon(self.icon_nothing)
        if pos_0[2] == 1:
            self.flag_pos_0_3 = 1
            self.pos_0_3.setIcon(self.icon_match)
        if pos_0[2] == 0:
            self.flag_pos_0_3 = 0
            self.pos_0_3.setIcon(self.icon_nothing)
        if pos_0[3] == 1:
            self.flag_pos_0_4 = 1
            self.pos_0_4.setIcon(self.icon_match_heng)
        if pos_0[3] == 0:
            self.flag_pos_0_4 = 0
            self.pos_0_4.setIcon(self.icon_nothing)
        if pos_0[4] == 1:
            self.flag_pos_0_5 = 1
            self.pos_0_5.setIcon(self.icon_match)
        if pos_0[4] == 0:
            self.flag_pos_0_5 = 0
            self.pos_0_5.setIcon(self.icon_nothing)
        if pos_0[5] == 1:
            self.flag_pos_0_6 = 1
            self.pos_0_6.setIcon(self.icon_match)
        if pos_0[5] == 0:
            self.flag_pos_0_6 = 0
            self.pos_0_6.setIcon(self.icon_nothing)
        if pos_0[6] == 1:
            self.flag_pos_0_7 = 1
            self.pos_0_7.setIcon(self.icon_match_heng)
        if pos_0[6] == 0:
            self.flag_pos_0_7 = 0
            self.pos_0_7.setIcon(self.icon_nothing)
        #   第二个数字
        if pos_1[0] == 1:
            self.flag_pos_1_1 = 1
            self.pos_1_1.setIcon(self.icon_match_heng)
        if pos_1[0] == 0:
            self.flag_pos_1_1 = 0
            self.pos_1_1.setIcon(self.icon_nothing)
        if pos_1[1] == 1:
            self.flag_pos_1_2 = 1
            self.pos_1_2.setIcon(self.icon_match)
        if pos_1[1] == 0:
            self.flag_pos_1_2 = 0
            self.pos_1_2.setIcon(self.icon_nothing)
        if pos_1[2] == 1:
            self.flag_pos_1_3 = 1
            self.pos_1_3.setIcon(self.icon_match)
        if pos_1[2] == 0:
            self.flag_pos_1_3 = 0
            self.pos_1_3.setIcon(self.icon_nothing)
        if pos_1[3] == 1:
            self.flag_pos_1_4 = 1
            self.pos_1_4.setIcon(self.icon_match_heng)
        if pos_1[3] == 0:
            self.flag_pos_1_4 = 0
            self.pos_1_4.setIcon(self.icon_nothing)
        if pos_1[4] == 1:
            self.flag_pos_1_5 = 1
            self.pos_1_5.setIcon(self.icon_match)
        if pos_1[4] == 0:
            self.flag_pos_1_5 = 0
            self.pos_1_5.setIcon(self.icon_nothing)
        if pos_1[5] == 1:
            self.flag_pos_1_6 = 1
            self.pos_1_6.setIcon(self.icon_match)
        if pos_1[5] == 0:
            self.flag_pos_1_6 = 0
            self.pos_1_6.setIcon(self.icon_nothing)
        if pos_1[6] == 1:
            self.flag_pos_1_7 = 1
            self.pos_1_7.setIcon(self.icon_match_heng)
        if pos_1[6] == 0:
            self.flag_pos_1_7 = 0
            self.pos_1_7.setIcon(self.icon_nothing)
        #   符号
        if pos_2[0] == 1:
            self.flag_sig_1 = 1
            self.sig_0.setIcon(self.icon_match_heng)
        if pos_2[0] == 0:
            self.flag_sig_1 = 0
            self.sig_0.setIcon(self.icon_nothing)
        if pos_2[1] == 1:
            self.flag_sig_2 = 1
            self.sig_1.setIcon(self.icon_match)
        if pos_2[1] == 0:
            self.flag_sig_2 = 0
            self.sig_1.setIcon(self.icon_nothing)
        if pos_2[2] == 1:
            self.flag_sig_3 = 1
        if pos_2[2] == 0:
            self.flag_sig_3 = 0
        if pos_2[3] == 1:
            self.flag_sig_4 = 1
        if pos_2[3] == 0:
            self.flag_sig_4 = 0
        if (self.flag_sig_3 == 1) & (self.flag_sig_4 == 1):
            self.sig_3_4.setText(QtCore.QCoreApplication.translate("MainWindow", "X"))
        if (pos_2[3] == 0) & (pos_2[2] == 0):
            self.sig_3_4.setText("")
        #   第三个数字
        if pos_3[0] == 1:
            self.flag_pos_3_1 = 1
            self.pos_3_1.setIcon(self.icon_match_heng)
        if pos_3[0] == 0:
            self.flag_pos_3_1 = 0
            self.pos_3_1.setIcon(self.icon_nothing)
        if pos_3[1] == 1:
            self.flag_pos_3_2 = 1
            self.pos_3_2.setIcon(self.icon_match)
        if pos_3[1] == 0:
            self.flag_pos_3_2 = 0
            self.pos_3_2.setIcon(self.icon_nothing)
        if pos_3[2] == 1:
            self.flag_pos_3_3 = 1
            self.pos_3_3.setIcon(self.icon_match)
        if pos_3[2] == 0:
            self.flag_pos_3_3 = 0
            self.pos_3_3.setIcon(self.icon_nothing)
        if pos_3[3] == 1:
            self.flag_pos_3_4 = 1
            self.pos_3_4.setIcon(self.icon_match_heng)
        if pos_3[3] == 0:
            self.flag_pos_3_4 = 0
            self.pos_3_4.setIcon(self.icon_nothing)
        if pos_3[4] == 1:
            self.flag_pos_3_5 = 1
            self.pos_3_5.setIcon(self.icon_match)
        if pos_3[4] == 0:
            self.flag_pos_3_5 = 0
            self.pos_3_5.setIcon(self.icon_nothing)
        if pos_3[5] == 1:
            self.flag_pos_3_6 = 1
            self.pos_3_6.setIcon(self.icon_match)
        if pos_3[5] == 0:
            self.flag_pos_3_6 = 0
            self.pos_3_6.setIcon(self.icon_nothing)
        if pos_3[6] == 1:
            self.flag_pos_3_7 = 1
            self.pos_3_7.setIcon(self.icon_match_heng)
        if pos_3[6] == 0:
            self.flag_pos_3_7 = 0
            self.pos_3_7.setIcon(self.icon_nothing)
        #   第四个数字
        if pos_4[0] == 1:
            self.flag_pos_4_1 = 1
            self.pos_4_1.setIcon(self.icon_match_heng)
        if pos_4[0] == 0:
            self.flag_pos_4_1 = 0
            self.pos_4_1.setIcon(self.icon_nothing)
        if pos_4[1] == 1:
            self.flag_pos_4_2 = 1
            self.pos_4_2.setIcon(self.icon_match)
        if pos_4[1] == 0:
            self.flag_pos_4_2 = 0
            self.pos_4_2.setIcon(self.icon_nothing)
        if pos_4[2] == 1:
            self.flag_pos_4_3 = 1
            self.pos_4_3.setIcon(self.icon_match)
        if pos_4[2] == 0:
            self.flag_pos_4_3 = 0
            self.pos_4_3.setIcon(self.icon_nothing)
        if pos_4[3] == 1:
            self.flag_pos_4_4 = 1
            self.pos_4_4.setIcon(self.icon_match_heng)
        if pos_4[3] == 0:
            self.flag_pos_4_4 = 0
            self.pos_4_4.setIcon(self.icon_nothing)
        if pos_4[4] == 1:
            self.flag_pos_4_5 = 1
            self.pos_4_5.setIcon(self.icon_match)
        if pos_4[4] == 0:
            self.flag_pos_4_5 = 0
            self.pos_4_5.setIcon(self.icon_nothing)
        if pos_4[5] == 1:
            self.flag_pos_4_6 = 1
            self.pos_4_6.setIcon(self.icon_match)
        if pos_4[5] == 0:
            self.flag_pos_4_6 = 0
            self.pos_4_6.setIcon(self.icon_nothing)
        if pos_4[6] == 1:
            self.flag_pos_4_7 = 1
            self.pos_4_7.setIcon(self.icon_match_heng)
        if pos_4[6] == 0:
            self.flag_pos_4_7 = 0
            self.pos_4_7.setIcon(self.icon_nothing)
        #   第五个数字
        if pos_5[0] == 1:
            self.flag_pos_5_1 = 1
            self.pos_5_1.setIcon(self.icon_match_heng)
        if pos_5[0] == 0:
            self.flag_pos_5_1 = 0
            self.pos_5_1.setIcon(self.icon_nothing)
        if pos_5[1] == 1:
            self.flag_pos_5_2 = 1
            self.pos_5_2.setIcon(self.icon_match)
        if pos_5[1] == 0:
            self.flag_pos_5_2 = 0
            self.pos_5_2.setIcon(self.icon_nothing)
        if pos_5[2] == 1:
            self.flag_pos_5_3 = 1
            self.pos_5_3.setIcon(self.icon_match)
        if pos_5[2] == 0:
            self.flag_pos_5_3 = 0
            self.pos_5_3.setIcon(self.icon_nothing)
        if pos_5[3] == 1:
            self.flag_pos_5_4 = 1
            self.pos_5_4.setIcon(self.icon_match_heng)
        if pos_5[3] == 0:
            self.flag_pos_5_4 = 0
            self.pos_5_4.setIcon(self.icon_nothing)
        if pos_5[4] == 1:
            self.flag_pos_5_5 = 1
            self.pos_5_5.setIcon(self.icon_match)
        if pos_5[4] == 0:
            self.flag_pos_5_5 = 0
            self.pos_5_5.setIcon(self.icon_nothing)
        if pos_5[5] == 1:
            self.flag_pos_5_6 = 1
            self.pos_5_6.setIcon(self.icon_match)
        if pos_5[5] == 0:
            self.flag_pos_5_6 = 0
            self.pos_5_6.setIcon(self.icon_nothing)
        if pos_5[6] == 1:
            self.flag_pos_5_7 = 1
            self.pos_5_7.setIcon(self.icon_match_heng)
        if pos_5[6] == 0:
            self.flag_pos_5_7 = 0
            self.pos_5_7.setIcon(self.icon_nothing)
        #   第六个数字
        if pos_6[0] == 1:
            self.flag_pos_6_1 = 1
            self.pos_6_1.setIcon(self.icon_match_heng)
        if pos_6[0] == 0:
            self.flag_pos_6_1 = 0
            self.pos_6_1.setIcon(self.icon_nothing)
        if pos_6[1] == 1:
            self.flag_pos_6_2 = 1
            self.pos_6_2.setIcon(self.icon_match)
        if pos_6[1] == 0:
            self.flag_pos_6_2 = 0
            self.pos_6_2.setIcon(self.icon_nothing)
        if pos_6[2] == 1:
            self.flag_pos_6_3 = 1
            self.pos_6_3.setIcon(self.icon_match)
        if pos_6[2] == 0:
            self.flag_pos_6_3 = 0
            self.pos_6_3.setIcon(self.icon_nothing)
        if pos_6[3] == 1:
            self.flag_pos_6_4 = 1
            self.pos_6_4.setIcon(self.icon_match_heng)
        if pos_6[3] == 0:
            self.flag_pos_6_4 = 0
            self.pos_6_4.setIcon(self.icon_nothing)
        if pos_6[4] == 1:
            self.flag_pos_6_5 = 1
            self.pos_6_5.setIcon(self.icon_match)
        if pos_6[4] == 0:
            self.flag_pos_6_5 = 0
            self.pos_6_5.setIcon(self.icon_nothing)
        if pos_6[5] == 1:
            self.flag_pos_6_6 = 1
            self.pos_6_6.setIcon(self.icon_match)
        if pos_6[5] == 0:
            self.flag_pos_6_6 = 0
            self.pos_6_6.setIcon(self.icon_nothing)
        if pos_6[6] == 1:
            self.flag_pos_6_7 = 1
            self.pos_6_7.setIcon(self.icon_match_heng)
        if pos_6[6] == 0:
            self.flag_pos_6_7 = 0
            self.pos_6_7.setIcon(self.icon_nothing)
        #print(self.flag_pos_0_1,self.flag_pos_0_2,self.flag_pos_0_3,self.flag_pos_0_4,self.flag_pos_0_5,self.flag_pos_0_6,
        #      self.flag_pos_1_1,self.flag_pos_1_2,self.flag_pos_1_3,self.flag_pos_1_4,self.flag_pos_1_5,self.flag_pos_1_6,
        #      self.flag_sig_1,self.flag_sig_2,self.flag_sig_3,self.flag_sig_4,
        #      self.flag_pos_3_1,self.flag_pos_3_2,self.flag_pos_3_3,self.flag_pos_3_4,self.flag_pos_3_5,self.flag_pos_3_6)


    #   根据question_number的值输出题目的函数
    def output_question(self):
        equal_initial = [copy.deepcopy(mt.num_empty),copy.deepcopy(mt.num_empty),copy.deepcopy(mt.sig_empty),copy.deepcopy(mt.num_empty),copy.deepcopy(mt.num_empty),
                         copy.deepcopy(mt.num_empty),copy.deepcopy(mt.num_empty)]
        if self.mode_number == 1:
            eq_str = self.table_1.cell(self.question_number, 0).value
        else:
            eq_str = self.table_2.cell(self.question_number, 0).value
        mt.trans_string(eq_str, equal_initial)
        if (equal_initial[2][2] == 1) & (equal_initial[2][3] == 1):
            self.equal_2.setCurrentIndex(1)
        else:
            self.equal_2.setCurrentIndex(0)
        self.set_match(equal_initial)
        self.equal_initial = copy.deepcopy(equal_initial)
        return equal_initial
    #   根据图读取目前的等式的函数
    def read_equal(self):

        equal = [copy.deepcopy(mt.num_empty),copy.deepcopy(mt.num_empty),copy.deepcopy(mt.sig_empty),copy.deepcopy(mt.num_empty),copy.deepcopy(mt.num_empty),
                         copy.deepcopy(mt.num_empty),copy.deepcopy(mt.num_empty)]
        num_0 = equal[0]
        num_1 = equal[1]
        sig = equal[2]
        num_3 = equal[3]
        num_4 = equal[4]
        num_5 = equal[5]
        num_6 = equal[6]

        if self.flag_pos_0_1 == 1:
            num_0[0] = 1
        if self.flag_pos_0_1 == 0:
            num_0[0] = 0
        if self.flag_pos_0_2 == 1:
            num_0[1] = 1
        if self.flag_pos_0_2 == 0:
            num_0[1] = 0
        if self.flag_pos_0_3 == 1:
            num_0[2] = 1
        if self.flag_pos_0_3 == 0:
            num_0[2] = 0
        if self.flag_pos_0_4 == 1:
            num_0[3] = 1
        if self.flag_pos_0_4 == 0:
            num_0[3] = 0
        if self.flag_pos_0_5 == 1:
            num_0[4] = 1
        if self.flag_pos_0_5 == 0:
            num_0[4] = 0
        if self.flag_pos_0_6 == 1:
            num_0[5] = 1
        if self.flag_pos_0_6 == 0:
            num_0[5] = 0
        if self.flag_pos_0_7 == 1:
            num_0[6] = 1
        if self.flag_pos_0_7 == 0:
            num_0[6] = 0
        if self.flag_pos_1_1 == 1:
            num_1[0] = 1
        if self.flag_pos_1_1 == 0:
            num_1[0] = 0
        if self.flag_pos_1_2 == 1:
            num_1[1] = 1
        if self.flag_pos_1_2 == 0:
            num_1[1] = 0
        if self.flag_pos_1_3 == 1:
            num_1[2] = 1
        if self.flag_pos_1_3 == 0:
            num_1[2] = 0
        if self.flag_pos_1_4 == 1:
            num_1[3] = 1
        if self.flag_pos_1_4 == 0:
            num_1[3] = 0
        if self.flag_pos_1_5 == 1:
            num_1[4] = 1
        if self.flag_pos_1_5 == 0:
            num_1[4] = 0
        if self.flag_pos_1_6 == 1:
            num_1[5] = 1
        if self.flag_pos_1_6 == 0:
            num_1[5] = 0
        if self.flag_pos_1_7 == 1:
            num_1[6] = 1
        if self.flag_pos_1_7 == 0:
            num_1[6] = 0
        if self.flag_sig_1 == 1:
            sig[0] = 1
        if self.flag_sig_1 == 0:
            sig[0] = 0
        if self.flag_sig_2 == 1:
            sig[1] = 1
        if self.flag_sig_2 == 0:
            sig[1] = 0
        if self.flag_sig_3 == 1:
            sig[2] = 1
        if self.flag_sig_3 == 0:
            sig[2] = 0
        if self.flag_sig_4 == 1:
            sig[3] = 1
        if self.flag_sig_4 == 0:
            sig[3] = 0
        if self.flag_pos_3_1 == 1:
            num_3[0] = 1
        if self.flag_pos_3_1 == 0:
            num_3[0] = 0
        if self.flag_pos_3_2 == 1:
            num_3[1] = 1
        if self.flag_pos_3_2 == 0:
            num_3[1] = 0
        if self.flag_pos_3_3 == 1:
            num_3[2] = 1
        if self.flag_pos_3_3 == 0:
            num_3[2] = 0
        if self.flag_pos_3_4 == 1:
            num_3[3] = 1
        if self.flag_pos_3_4 == 0:
            num_3[3] = 0
        if self.flag_pos_3_5 == 1:
            num_3[4] = 1
        if self.flag_pos_3_5 == 0:
            num_3[4] = 0
        if self.flag_pos_3_6 == 1:
            num_3[5] = 1
        if self.flag_pos_3_6 == 0:
            num_3[5] = 0
        if self.flag_pos_3_7 == 1:
            num_3[6] = 1
        if self.flag_pos_3_7 == 0:
            num_3[6] = 0
        if self.flag_pos_4_1 == 1:
            num_4[0] = 1
        if self.flag_pos_4_1 == 0:
            num_4[0] = 0
        if self.flag_pos_4_2 == 1:
            num_4[1] = 1
        if self.flag_pos_4_2 == 0:
            num_4[1] = 0
        if self.flag_pos_4_3 == 1:
            num_4[2] = 1
        if self.flag_pos_4_3 == 0:
            num_4[2] = 0
        if self.flag_pos_4_4 == 1:
            num_4[3] = 1
        if self.flag_pos_4_4 == 0:
            num_4[3] = 0
        if self.flag_pos_4_5 == 1:
            num_4[4] = 1
        if self.flag_pos_4_5 == 0:
            num_4[4] = 0
        if self.flag_pos_4_6 == 1:
            num_4[5] = 1
        if self.flag_pos_4_6 == 0:
            num_4[5] = 0
        if self.flag_pos_4_7 == 1:
            num_4[6] = 1
        if self.flag_pos_4_7 == 0:
            num_4[6] = 0
        if self.flag_pos_5_1 == 1:
            num_5[0] = 1
        else:
            num_5[0] = 0
        if self.flag_pos_5_2 == 1:
            num_5[1] = 1
        else:
            num_5[1] = 0
        if self.flag_pos_5_3 == 1:
            num_5[2] = 1
        else:
            num_5[2] = 0
        if self.flag_pos_5_4 == 1:
            num_5[3] = 1
        else:
            num_5[3] = 0
        if self.flag_pos_5_5 == 1:
            num_5[4] = 1
        else:
            num_5[4] = 0
        if self.flag_pos_5_6 == 1:
            num_5[5] = 1
        else:
            num_5[5] = 0
        if self.flag_pos_5_7 == 1:
            num_5[6] = 1
        else:
            num_5[6] = 0
        if self.flag_pos_6_1 == 1:
            num_6[0] = 1
        else:
            num_6[0] = 0
        if self.flag_pos_6_2 == 1:
            num_6[1] = 1
        else:
            num_6[1] = 0
        if self.flag_pos_6_3 == 1:
            num_6[2] = 1
        else:
            num_6[2] = 0
        if self.flag_pos_6_4 == 1:
            num_6[3] = 1
        else:
            num_6[3] = 0
        if self.flag_pos_6_5 == 1:
            num_6[4] = 1
        else:
            num_6[4] = 0
        if self.flag_pos_6_6 == 1:
            num_6[5] = 1
        else:
            num_6[5] = 0
        if self.flag_pos_6_7 == 1:
            num_6[6] = 1
        else:
            num_6[6] = 0

        return equal

    #   判断移动了几根的函数
    def sticks_moved(self):
        equal_initial  = copy.deepcopy(self.equal_initial)
        equal = self.read_equal()
        print(equal_initial)
        print(equal)
        count = 0
        for i in range(7):
            if equal_initial[0][i]!=equal[0][i]:
                count += 1
            if equal_initial[1][i]!=equal[1][i]:
                count += 1
            if equal_initial[3][i]!=equal[3][i]:
                count += 1
            if equal_initial[4][i]!=equal[4][i]:
                count += 1
            if equal_initial[5][i]!=equal[5][i]:
                count += 1
            if equal_initial[6][i]!=equal[6][i]:
                count += 1
        for i in range(4):
            if equal_initial[2][i]!=equal[2][i]:
                count += 1
        return count

    #   得到更多等式的函数
    def produce_question(self):
        #   随机生成等式：随机生成第一个数，第二个数，一个1~3之间的数然后得到符号，结果数，中间添加等号，然后判断
        while True:
            #   循环，直到生成新的等式为止
            random_str = ''
            num_1 = random.randint(1,99)
            num_2 = random.randint(1,99)
            sig_num = random.randint(1,2)
            num_3 = random.randint(1,99)
            random_str += str(num_1)
            if sig_num == 1:
                random_str += '+'
            if sig_num == 2:
                random_str += '-'
            random_str += str(num_2)

            random_str += '='
            random_str += str(num_3)
            #   至此生成了一个等式字符串
            #   检查等式库中是否存在
            flag_existed = 0
            for i in range(self.question_number_mode1):
                if random_str == self.worksheet_1.cell(i+2, 1).value:
                    flag_existed = 1
            #   等式库中不存在
            if flag_existed == 0:
                #   判断等式是否成立
                equal = [copy.deepcopy(mt.num_empty), copy.deepcopy(mt.num_empty), copy.deepcopy(mt.sig_empty),
                         copy.deepcopy(mt.num_empty), copy.deepcopy(mt.num_empty),
                         copy.deepcopy(mt.num_empty), copy.deepcopy(mt.num_empty)]
                if mt.trans_string(random_str, equal) == 1:
                    outcome, equal = search.search_solution(equal)
                    #   等式成立
                    if outcome == 1:
                        #   写入
                        self.worksheet_1.cell(self.question_number_mode1 + 2, 1, random_str)
                        self.worksheet_1.cell(self.question_number_mode1 + 2, 2, mt.equal2str(equal))
                        self.workbook_1.save('library_1.xlsx')
                        #   更新值
                        self.question_number_mode1 += 1
                        self.data_1 = xlrd.open_workbook('library_1.xlsx')
                        self.workbook_1 = openpyxl.load_workbook('library_1.xlsx')
                        self.data_2 = xlrd.open_workbook('library_2.xlsx')
                        self.workbook_2 = openpyxl.load_workbook('library_2.xlsx')
                        self.table_1 = self.data_1.sheets()[0]
                        self.worksheet_1 = self.workbook_1.worksheets[0]
                        self.worksheet_2 = self.workbook_2.worksheets[0]
                        self.table_2 = self.data_2.sheets()[0]
                        #   显示
                        self.question_number = self.question_number_mode1
                        self.mode_number = 1
                        self.match_movable = 1
                        self.output_question()
                        self.showanswer_button.setEnabled(True)
                        self.submit_button.setEnabled(True)
                        return 1


        #   生成等式的条件：可解+等式库中不存在
        #   生成等式后，要将等式：1.储存在excel中；2.显示在游戏界面中
    #   生成等式的函数
    def equal_generate_clicked(self):
        #   随机生成等式：随机生成第一个数，第二个数，一个1~3之间的数然后得到符号，结果数，中间添加等号，然后判断
        while True:
            #   循环，直到生成新的等式为止
            random_str = ''
            num_1 = random.randint(1,99)
            num_2 = random.randint(1,99)
            sig_num = random.randint(1,2)
            num_3 = random.randint(1,99)
            random_str += str(num_1)
            if sig_num == 1:
                random_str += '+'
            if sig_num == 2:
                random_str += '-'
            random_str += str(num_2)

            random_str += '='
            random_str += str(num_3)
            #   至此生成了一个等式字符串
            #   检查等式库中是否存在
            flag_existed = 0
            for i in range(self.question_number_mode1):
                if random_str == self.worksheet_1.cell(i+2, 1).value:
                    flag_existed = 1
            #   等式库中不存在
            if flag_existed == 0:
                #   判断等式是否成立
                equal = [copy.deepcopy(mt.num_empty), copy.deepcopy(mt.num_empty), copy.deepcopy(mt.sig_empty),
                         copy.deepcopy(mt.num_empty), copy.deepcopy(mt.num_empty),
                         copy.deepcopy(mt.num_empty), copy.deepcopy(mt.num_empty)]
                if mt.trans_string(random_str, equal) == 1:
                    if mt.judge_equal(equal) == 1:
                        outcome, equal, count = search.equal_solution(equal)
                        #   等式成立
                        if outcome == 1:
                            #   提示
                            self.level.label.setText(QtCore.QCoreApplication.translate("Dialog","生成等式成功！难度："+str(count)    ))
                            self.level.show()
                            #   写入
                            self.worksheet_1.cell(self.question_number_mode1 + 2, 1, random_str)
                            self.worksheet_1.cell(self.question_number_mode1 + 2, 2, mt.equal2str(equal))
                            self.workbook_1.save('library_1.xlsx')
                            #   更新值
                            self.question_number_mode1 += 1
                            self.data_1 = xlrd.open_workbook('library_1.xlsx')
                            self.workbook_1 = openpyxl.load_workbook('library_1.xlsx')
                            self.data_2 = xlrd.open_workbook('library_2.xlsx')
                            self.workbook_2 = openpyxl.load_workbook('library_2.xlsx')
                            self.table_1 = self.data_1.sheets()[0]
                            self.worksheet_1 = self.workbook_1.worksheets[0]
                            self.worksheet_2 = self.workbook_2.worksheets[0]
                            self.table_2 = self.data_2.sheets()[0]
                            #   显示
                            self.mode_number = 1
                            self.match_movable = 1
                            self.question_number = self.question_number_mode1
                            self.output_question()
                            self.showanswer_button.setEnabled(True)
                            self.submit_button.setEnabled(True)
                            #   把按键全都解封
                            self.pos_0_1.setEnabled(True)
                            self.pos_0_2.setEnabled(True)
                            self.pos_0_3.setEnabled(True)
                            self.pos_0_4.setEnabled(True)
                            self.pos_0_5.setEnabled(True)
                            self.pos_0_6.setEnabled(True)
                            self.pos_0_7.setEnabled(True)
                            self.pos_1_1.setEnabled(True)
                            self.pos_1_2.setEnabled(True)
                            self.pos_1_3.setEnabled(True)
                            self.pos_1_4.setEnabled(True)
                            self.pos_1_5.setEnabled(True)
                            self.pos_1_6.setEnabled(True)
                            self.pos_1_7.setEnabled(True)
                            self.pos_3_1.setEnabled(True)
                            self.pos_3_2.setEnabled(True)
                            self.pos_3_3.setEnabled(True)
                            self.pos_3_4.setEnabled(True)
                            self.pos_3_5.setEnabled(True)
                            self.pos_3_6.setEnabled(True)
                            self.pos_3_7.setEnabled(True)
                            self.pos_4_1.setEnabled(True)
                            self.pos_4_2.setEnabled(True)
                            self.pos_4_3.setEnabled(True)
                            self.pos_4_4.setEnabled(True)
                            self.pos_4_5.setEnabled(True)
                            self.pos_4_6.setEnabled(True)
                            self.pos_4_7.setEnabled(True)
                            self.pos_5_1.setEnabled(True)
                            self.pos_5_2.setEnabled(True)
                            self.pos_5_3.setEnabled(True)
                            self.pos_5_4.setEnabled(True)
                            self.pos_5_5.setEnabled(True)
                            self.pos_5_6.setEnabled(True)
                            self.pos_5_7.setEnabled(True)
                            self.pos_6_1.setEnabled(True)
                            self.pos_6_2.setEnabled(True)
                            self.pos_6_3.setEnabled(True)
                            self.pos_6_4.setEnabled(True)
                            self.pos_6_5.setEnabled(True)
                            self.pos_6_6.setEnabled(True)
                            self.pos_6_7.setEnabled(True)
                            self.sig_0.setEnabled(True)
                            self.sig_1.setEnabled(True)
                            self.sig_3_4.setEnabled(True)
                            self.pushButton_59.setEnabled(True)
                            self.pushButton_60.setEnabled(True)
                            self.submit_button.setEnabled(True)
                            self.showanswer_button.setEnabled(True)
                            return 1
    #
    #   定义一个点击后火柴棍消失或出现的函数
    #   1,4,7为match_heng
    def match_pos_0_1(self):
        # 如果火柴棍点亮
        if self.flag_pos_0_1 == 1:
            self.pos_0_1.setIcon(self.icon_nothing)
            self.flag_pos_0_1 = 0
        #  未点亮则点亮
        else:
            self.pos_0_1.setIcon(self.icon_match_heng)
            self.flag_pos_0_1 = 1
    def match_pos_0_2(self):
        # 如果火柴棍点亮
        if self.flag_pos_0_2 == 1:
            self.pos_0_2.setIcon(self.icon_nothing)
            self.flag_pos_0_2 = 0
        #  未点亮则点亮
        else:
            self.pos_0_2.setIcon(self.icon_match)
            self.flag_pos_0_2 = 1
    def match_pos_0_3(self):
        # 如果火柴棍点亮
        if self.flag_pos_0_3 == 1:
            self.pos_0_3.setIcon(self.icon_nothing)
            self.flag_pos_0_3 = 0
        #  未点亮则点亮
        else:
            self.pos_0_3.setIcon(self.icon_match)
            self.flag_pos_0_3 = 1
    def match_pos_0_4(self):
        # 如果火柴棍点亮
        if self.flag_pos_0_4 == 1:
            self.pos_0_4.setIcon(self.icon_nothing)
            self.flag_pos_0_4 = 0
        #  未点亮则点亮
        else:
            self.pos_0_4.setIcon(self.icon_match_heng)
            self.flag_pos_0_4 = 1
    def match_pos_0_5(self):
        # 如果火柴棍点亮
        if self.flag_pos_0_5 == 1:
            self.pos_0_5.setIcon(self.icon_nothing)
            self.flag_pos_0_5 = 0
        #  未点亮则点亮
        else:
            self.pos_0_5.setIcon(self.icon_match)
            self.flag_pos_0_5 = 1
    def match_pos_0_6(self):
        # 如果火柴棍点亮
        if self.flag_pos_0_6 == 1:
            self.pos_0_6.setIcon(self.icon_nothing)
            self.flag_pos_0_6 = 0
        #  未点亮则点亮
        else:
            self.pos_0_6.setIcon(self.icon_match)
            self.flag_pos_0_6 = 1
    def match_pos_0_7(self):
        # 如果火柴棍点亮
        if self.flag_pos_0_7 == 1:
            self.pos_0_7.setIcon(self.icon_nothing)
            self.flag_pos_0_7 = 0
        #  未点亮则点亮
        else:
            self.pos_0_7.setIcon(self.icon_match_heng)
            self.flag_pos_0_7 = 1
    def match_pos_1_1(self):
        # 如果火柴棍点亮
        if self.flag_pos_1_1 == 1:
            self.pos_1_1.setIcon(self.icon_nothing)
            self.flag_pos_1_1 = 0
        #  未点亮则点亮
        else:
            self.pos_1_1.setIcon(self.icon_match_heng)
            self.flag_pos_1_1 = 1
    def match_pos_1_2(self):
        # 如果火柴棍点亮
        if self.flag_pos_1_2 == 1:
            self.pos_1_2.setIcon(self.icon_nothing)
            self.flag_pos_1_2 = 0
        #  未点亮则点亮
        else:
            self.pos_1_2.setIcon(self.icon_match)
            self.flag_pos_1_2 = 1
    def match_pos_1_3(self):
        # 如果火柴棍点亮
        if self.flag_pos_1_3 == 1:
            self.pos_1_3.setIcon(self.icon_nothing)
            self.flag_pos_1_3 = 0
        #  未点亮则点亮
        else:
            self.pos_1_3.setIcon(self.icon_match)
            self.flag_pos_1_3 = 1
    def match_pos_1_4(self):
        # 如果火柴棍点亮
        if self.flag_pos_1_4 == 1:
            self.pos_1_4.setIcon(self.icon_nothing)
            self.flag_pos_1_4 = 0
        #  未点亮则点亮
        else:
            self.pos_1_4.setIcon(self.icon_match_heng)
            self.flag_pos_1_4 = 1
    def match_pos_1_5(self):
        # 如果火柴棍点亮
        if self.flag_pos_1_5 == 1:
            self.pos_1_5.setIcon(self.icon_nothing)
            self.flag_pos_1_5 = 0
        #  未点亮则点亮
        else:
            self.pos_1_5.setIcon(self.icon_match)
            self.flag_pos_1_5 = 1
    def match_pos_1_6(self):
        # 如果火柴棍点亮
        if self.flag_pos_1_6 == 1:
            self.pos_1_6.setIcon(self.icon_nothing)
            self.flag_pos_1_6 = 0
        #  未点亮则点亮
        else:
            self.pos_1_6.setIcon(self.icon_match)
            self.flag_pos_1_6 = 1
    def match_pos_1_7(self):
        # 如果火柴棍点亮
        if self.flag_pos_1_7 == 1:
            self.pos_1_7.setIcon(self.icon_nothing)
            self.flag_pos_1_7 = 0
        #  未点亮则点亮
        else:
            self.pos_1_7.setIcon(self.icon_match_heng)
            self.flag_pos_1_7 = 1
    def match_pos_3_1(self):
        # 如果火柴棍点亮
        if self.flag_pos_3_1 == 1:
            self.pos_3_1.setIcon(self.icon_nothing)
            self.flag_pos_3_1 = 0
        #  未点亮则点亮
        else:
            self.pos_3_1.setIcon(self.icon_match_heng)
            self.flag_pos_3_1 = 1
    def match_pos_3_2(self):
        # 如果火柴棍点亮
        if self.flag_pos_3_2 == 1:
            self.pos_3_2.setIcon(self.icon_nothing)
            self.flag_pos_3_2 = 0
        #  未点亮则点亮
        else:
            self.pos_3_2.setIcon(self.icon_match)
            self.flag_pos_3_2 = 1
    def match_pos_3_3(self):
        # 如果火柴棍点亮
        if self.flag_pos_3_3 == 1:
            self.pos_3_3.setIcon(self.icon_nothing)
            self.flag_pos_3_3 = 0
        #  未点亮则点亮
        else:
            self.pos_3_3.setIcon(self.icon_match)
            self.flag_pos_3_3 = 1
    def match_pos_3_4(self):
        # 如果火柴棍点亮
        if self.flag_pos_3_4 == 1:
            self.pos_3_4.setIcon(self.icon_nothing)
            self.flag_pos_3_4 = 0
        #  未点亮则点亮
        else:
            self.pos_3_4.setIcon(self.icon_match_heng)
            self.flag_pos_3_4 = 1
    def match_pos_3_5(self):
        # 如果火柴棍点亮
        if self.flag_pos_3_5 == 1:
            self.pos_3_5.setIcon(self.icon_nothing)
            self.flag_pos_3_5 = 0
        #  未点亮则点亮
        else:
            self.pos_3_5.setIcon(self.icon_match)
            self.flag_pos_3_5 = 1
    def match_pos_3_6(self):
        # 如果火柴棍点亮
        if self.flag_pos_3_6 == 1:
            self.pos_3_6.setIcon(self.icon_nothing)
            self.flag_pos_3_6 = 0
        #  未点亮则点亮
        else:
            self.pos_3_6.setIcon(self.icon_match)
            self.flag_pos_3_6 = 1
    def match_pos_3_7(self):
        # 如果火柴棍点亮
        if self.flag_pos_3_7 == 1:
            self.pos_3_7.setIcon(self.icon_nothing)
            self.flag_pos_3_7 = 0
        #  未点亮则点亮
        else:
            self.pos_3_7.setIcon(self.icon_match_heng)
            self.flag_pos_3_7 = 1
    def match_pos_4_1(self):
        # 如果火柴棍点亮
        if self.flag_pos_4_1 == 1:
            self.pos_4_1.setIcon(self.icon_nothing)
            self.flag_pos_4_1 = 0
        #  未点亮则点亮
        else:
            self.pos_4_1.setIcon(self.icon_match_heng)
            self.flag_pos_4_1 = 1
    def match_pos_4_2(self):
        # 如果火柴棍点亮
        if self.flag_pos_4_2 == 1:
            self.pos_4_2.setIcon(self.icon_nothing)
            self.flag_pos_4_2 = 0
        #  未点亮则点亮
        else:
            self.pos_4_2.setIcon(self.icon_match)
            self.flag_pos_4_2 = 1
    def match_pos_4_3(self):
        # 如果火柴棍点亮
        if self.flag_pos_4_3 == 1:
            self.pos_4_3.setIcon(self.icon_nothing)
            self.flag_pos_4_3 = 0
        #  未点亮则点亮
        else:
            self.pos_4_3.setIcon(self.icon_match)
            self.flag_pos_4_3 = 1
    def match_pos_4_4(self):
        # 如果火柴棍点亮
        if self.flag_pos_4_4 == 1:
            self.pos_4_4.setIcon(self.icon_nothing)
            self.flag_pos_4_4 = 0
        #  未点亮则点亮
        else:
            self.pos_4_4.setIcon(self.icon_match_heng)
            self.flag_pos_4_4 = 1
    def match_pos_4_5(self):
        # 如果火柴棍点亮
        if self.flag_pos_4_5 == 1:
            self.pos_4_5.setIcon(self.icon_nothing)
            self.flag_pos_4_5 = 0
        #  未点亮则点亮
        else:
            self.pos_4_5.setIcon(self.icon_match)
            self.flag_pos_4_5 = 1
    def match_pos_4_6(self):
        # 如果火柴棍点亮
        if self.flag_pos_4_6 == 1:
            self.pos_4_6.setIcon(self.icon_nothing)
            self.flag_pos_4_6 = 0
        #  未点亮则点亮
        else:
            self.pos_4_6.setIcon(self.icon_match)
            self.flag_pos_4_6 = 1
    def match_pos_4_7(self):
        # 如果火柴棍点亮
        if self.flag_pos_4_7 == 1:
            self.pos_4_7.setIcon(self.icon_nothing)
            self.flag_pos_4_7 = 0
        #  未点亮则点亮
        else:
            self.pos_4_7.setIcon(self.icon_match_heng)
            self.flag_pos_4_7 = 1
    def match_pos_5_1(self):
        # 如果火柴棍点亮
        if self.flag_pos_5_1 == 1:
            self.pos_5_1.setIcon(self.icon_nothing)
            self.flag_pos_5_1 = 0
        #  未点亮则点亮
        else:
            self.pos_5_1.setIcon(self.icon_match_heng)
            self.flag_pos_5_1 = 1
    def match_pos_5_2(self):
        # 如果火柴棍点亮
        if self.flag_pos_5_2 == 1:
            self.pos_5_2.setIcon(self.icon_nothing)
            self.flag_pos_5_2 = 0
        #  未点亮则点亮
        else:
            self.pos_5_2.setIcon(self.icon_match)
            self.flag_pos_5_2 = 1
    def match_pos_5_3(self):
        # 如果火柴棍点亮
        if self.flag_pos_5_3 == 1:
            self.pos_5_3.setIcon(self.icon_nothing)
            self.flag_pos_5_3 = 0
        #  未点亮则点亮
        else:
            self.pos_5_3.setIcon(self.icon_match)
            self.flag_pos_5_3 = 1
    def match_pos_5_4(self):
        # 如果火柴棍点亮
        if self.flag_pos_5_4 == 1:
            self.pos_5_4.setIcon(self.icon_nothing)
            self.flag_pos_5_4 = 0
        #  未点亮则点亮
        else:
            self.pos_5_4.setIcon(self.icon_match_heng)
            self.flag_pos_5_4 = 1
    def match_pos_5_5(self):
        # 如果火柴棍点亮
        if self.flag_pos_5_5 == 1:
            self.pos_5_5.setIcon(self.icon_nothing)
            self.flag_pos_5_5 = 0
        #  未点亮则点亮
        else:
            self.pos_5_5.setIcon(self.icon_match)
            self.flag_pos_5_5 = 1
    def match_pos_5_6(self):
        # 如果火柴棍点亮
        if self.flag_pos_5_6 == 1:
            self.pos_5_6.setIcon(self.icon_nothing)
            self.flag_pos_5_6 = 0
        #  未点亮则点亮
        else:
            self.pos_5_6.setIcon(self.icon_match)
            self.flag_pos_5_6 = 1
    def match_pos_5_7(self):
        # 如果火柴棍点亮
        if self.flag_pos_5_7 == 1:
            self.pos_5_7.setIcon(self.icon_nothing)
            self.flag_pos_5_7 = 0
        #  未点亮则点亮
        else:
            self.pos_5_7.setIcon(self.icon_match_heng)
            self.flag_pos_5_7 = 1
    def match_pos_6_1(self):
        # 如果火柴棍点亮
        if self.flag_pos_6_1 == 1:
            self.pos_6_1.setIcon(self.icon_nothing)
            self.flag_pos_6_1 = 0
        #  未点亮则点亮
        else:
            self.pos_6_1.setIcon(self.icon_match_heng)
            self.flag_pos_6_1 = 1
    def match_pos_6_2(self):
        # 如果火柴棍点亮
        if self.flag_pos_6_2 == 1:
            self.pos_6_2.setIcon(self.icon_nothing)
            self.flag_pos_6_2 = 0
        #  未点亮则点亮
        else:
            self.pos_6_2.setIcon(self.icon_match)
            self.flag_pos_6_2 = 1
    def match_pos_6_3(self):
        # 如果火柴棍点亮
        if self.flag_pos_6_3 == 1:
            self.pos_6_3.setIcon(self.icon_nothing)
            self.flag_pos_6_3 = 0
        #  未点亮则点亮
        else:
            self.pos_6_3.setIcon(self.icon_match)
            self.flag_pos_6_3 = 1
    def match_pos_6_4(self):
        # 如果火柴棍点亮
        if self.flag_pos_6_4 == 1:
            self.pos_6_4.setIcon(self.icon_nothing)
            self.flag_pos_6_4 = 0
        #  未点亮则点亮
        else:
            self.pos_6_4.setIcon(self.icon_match_heng)
            self.flag_pos_6_4 = 1
    def match_pos_6_5(self):
        # 如果火柴棍点亮
        if self.flag_pos_6_5 == 1:
            self.pos_6_5.setIcon(self.icon_nothing)
            self.flag_pos_6_5 = 0
        #  未点亮则点亮
        else:
            self.pos_6_5.setIcon(self.icon_match)
            self.flag_pos_6_5 = 1
    def match_pos_6_6(self):
        # 如果火柴棍点亮
        if self.flag_pos_6_6 == 1:
            self.pos_6_6.setIcon(self.icon_nothing)
            self.flag_pos_6_6 = 0
        #  未点亮则点亮
        else:
            self.pos_6_6.setIcon(self.icon_match)
            self.flag_pos_6_6 = 1
    def match_pos_6_7(self):
        # 如果火柴棍点亮
        if self.flag_pos_6_7 == 1:
            self.pos_6_7.setIcon(self.icon_nothing)
            self.flag_pos_6_7 = 0
        #  未点亮则点亮
        else:
            self.pos_6_7.setIcon(self.icon_match_heng)
            self.flag_pos_6_7 = 1
    #   下面几个是关于运算符号点击的函数
    #   加号的两根火柴棍点击
    def match_sig_1(self):
        if self.flag_sig_1 == 1:
            self.sig_0.setIcon(self.icon_nothing)
            self.flag_sig_1 = 0
        else:
            self.sig_0.setIcon(self.icon_match_heng)
            self.flag_sig_1 = 1
    def match_sig_2(self):
        if self.flag_sig_2 == 1:
            self.sig_1.setIcon(self.icon_nothing)
            self.flag_sig_2 = 0
        else:
            self.sig_1.setIcon(self.icon_match)
            self.flag_sig_2 = 1
    #   乘号点击
    def match_sig_3_4(self):
        if (self.flag_sig_3 == 1) & (self.flag_sig_4 == 1):
            self.sig_3_4.setText("")
            self.flag_sig_3 = 0
            self.flag_sig_4 = 0
        else:
            self.sig_3_4.setText(QtCore.QCoreApplication.translate("MainWindow", "X"))
            self.flag_sig_3 = 1
            self.flag_sig_4 = 1

#   选择关卡对话框
class question_dialog(QtWidgets.QWidget, question_dialog):
    def __init__(self):
        super(question_dialog, self).__init__()
        self.setupUi(self)

#   选择模式对话框
class mode_dialog(QtWidgets.QWidget, mode_dialog):
    def __init__(self):
        super(mode_dialog, self).__init__()
        self.setupUi(self)


#   判断回答正确错误的对话框
#   回答正确
class answerright_dialog(QtWidgets.QWidget, answerright_dialog):
    def __init__(self):
        super(answerright_dialog, self).__init__()
        self.setupUi(self)


#   回答错误
class answerwrong_dialog(QtWidgets.QWidget, answerwrong_dialog):
    def __init__(self):
        super(answerwrong_dialog, self).__init__()
        self.setupUi(self)


#   移动了太多火柴棍
class answermovetoomuch_dialog(QtWidgets.QWidget, answermovetoomuch_dialog):
    def __init__(self):
        super(answermovetoomuch_dialog, self).__init__()
        self.setupUi(self)

#   我要出题
class uploadquestion_dialog(QtWidgets.QWidget, uploadquestion_dialog):
    def __init__(self):
        super(uploadquestion_dialog, self).__init__()
        self.setupUi(self)

#   出题已存在
class upload_existed_dialog(QtWidgets.QWidget, upload_existed_dialog):
    def __init__(self):
        super(upload_existed_dialog, self).__init__()
        self.setupUi(self)
#   出题成功
class upload_success_dialog(QtWidgets.QWidget, upload_success_dialog):
    def __init__(self):
        super(upload_success_dialog, self).__init__()
        self.setupUi(self)

#
class produce_success_dialog(QtWidgets.QWidget, produce_success_dialog):
    def __init__(self):
        super(produce_success_dialog, self).__init__()
        self.setupUi(self)

#   出题无解
class upload_no_solution_dialog(QtWidgets.QWidget, upload_no_solution_dialog):
    def __init__(self):
        super(upload_no_solution_dialog, self).__init__()
        self.setupUi(self)

class level_dialog(QtWidgets.QWidget, level_dialog):
    def __init__(self):
        super(level_dialog, self).__init__()
        self.setupUi(self)


#   import match_rc
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MW = MainWindow()
    MW.show()
    sys.exit(app.exec_())